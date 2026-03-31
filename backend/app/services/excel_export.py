"""
스크리닝 결과 엑셀 내보내기 서비스
-- 수식을 포함하여 사용자가 직접 셀 데이터를 수정/검토 가능하도록 생성
-- Sheet 1: 종합 리포트 (가공된 지표)
-- Sheet 2: 산출 근거 (원본 데이터 + 엑셀 수식)
"""

import io
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter

def generate_screening_excel(top_picks: list) -> bytes:
    """
    스크리닝 결과를 엑셀 바이트 스트림으로 변환
    - 수식을 포함하는 Sheet 2 추가
    """
    output = io.BytesIO()
    
    # 1. 메인 리트용 DataFrame (표시용 데이터)
    report_data = []
    for p in top_picks:
        report_data.append({
            "종목명": p["name"],
            "종목코드": p["stock_code"],
            "현재가": p["price"],
            "기준일자": p["price_date"],
            "시가총액(억)": p["market_cap_billion"],
            "PER": p["per"],
            "PBR": p["pbr"],
            "ROE(%)": p["roe"],
            "PEG": p["peg"],
            "y-2 매출(억)": p.get("revenue_prev"),
            "y-2 영업이익(억)": p.get("operating_income_prev"),
            "y-1 매출(억)": p.get("revenue"),
            "y-1 영업이익(억)": p.get("operating_income"),
            "y-1 영업이익률(%)": p.get("operating_margin"),
            "매출성장률(%)": p["revenue_growth"],
            "영업이익성장률(%)": p["op_income_growth"],
            "재무효율": "✅" if p["financial_efficiency"] else "⚠️",
            "종합스코어": p["score"],
            "시그널": p["signal"]
        })
    df_report = pd.DataFrame(report_data)

    # 2. 산출 근거 데이터 (수식용 원본 데이터)
    # 여기서는 엑셀 내에서 직접 계산이 가능하도록 원본 원(won) 단위를 유지하거나 명시함
    calc_data = []
    for p in top_picks:
        # screening_engine.py의 분석 단계에서 원본 raw_data를 넘겨받아야 하지만, 
        # 현재는 가공된 값을 역산하거나 보관된 원본을 사용
        # (주의: 실제 원본 데이터 규모는 1e12 등 매우 큼)
        calc_data.append({
            "종목명": p["name"],
            "종목코드": p["stock_code"],
            "분석연도": p.get("bsns_year", "2025"),
            "시가총액(억)": p.get("market_cap_billion", 0),
            "당기순이익(억)": round((p.get("market_cap_billion", 0) / p["per"]), 2) if p.get("per") and p.get("market_cap_billion") else 0,
            "자본총계(억)": round((p.get("market_cap_billion", 0) / p["pbr"]), 2) if p.get("pbr") and p.get("market_cap_billion") else 0,
            "y-2 매출액(억)": p.get("revenue_prev", 0),
            "y-2 영업이익(억)": p.get("operating_income_prev", 0),
            "y-1 매출액(억)": p.get("revenue", 0),
            "y-1 영업이익(억)": p.get("operating_income", 0),
        })
    df_calc = pd.DataFrame(calc_data)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet 1: 종합 리포트
        df_report.to_excel(writer, sheet_name="종합 리포트", index=False)
        
        # Sheet 2: 산출 근거 및 수식
        # 수식 적용을 위해 원본 데이터를 먼저 쓰고, 그 옆에 수식 컬럼 추가
        df_calc.to_excel(writer, sheet_name="산출 근거", index=False)
        
        workbook = writer.book
        sheet2 = workbook["산출 근거"]
        
        # 헤더 추가 (수식 기반 지표)
        start_col = len(df_calc.columns) + 1
        headers = ["수식 PER", "수식 PBR", "수식 ROE(%)", "수식 영업이익률(%)", "수식 매출성장률(%)", "수식 영업이익성장률(%)"]
        for i, header in enumerate(headers):
            sheet2.cell(row=1, column=start_col + i, value=header)
            
        # 데이터 행마다 수식 삽입 
        # 열 매핑: A=명, B=코드, C=연도, D=시총, E=순이익, F=자본, G=y-2매출, H=y-2영업, I=y-1매출, J=y-1영업
        for row_idx in range(2, len(df_calc) + 2):
            sheet2.cell(row=row_idx, column=start_col, value=f"=D{row_idx}/E{row_idx}")
            sheet2.cell(row=row_idx, column=start_col + 1, value=f"=D{row_idx}/F{row_idx}")
            sheet2.cell(row=row_idx, column=start_col + 2, value=f"=E{row_idx}/F{row_idx}*100")
            sheet2.cell(row=row_idx, column=start_col + 3, value=f"=(J{row_idx}/I{row_idx})*100")
            # 매출성장률: (y-1 매출 - y-2 매출) / ABS(y-2 매출)
            sheet2.cell(row=row_idx, column=start_col + 4, value=f"=(I{row_idx}-G{row_idx})/ABS(G{row_idx})*100")
            # 영업이익성장률: (y-1 영업 - y-2 영업) / ABS(y-2 영업)
            sheet2.cell(row=row_idx, column=start_col + 5, value=f"=(J{row_idx}-H{row_idx})/ABS(H{row_idx})*100")

        # 시트 서식 조정 (열 너비 등)
        for sheet in [workbook["종합 리포트"], sheet2]:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

    return output.getvalue()
